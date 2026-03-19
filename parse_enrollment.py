"""
parse_enrollment.py
Reads the full enrollment history Excel file and returns structured
enrollment data for a given center and reference date.

Classification rules:
- First enrollment ever for a student = 'new'
- Gap > 60 days since last enrollment ended = 're-enrollment'
- Gap <= 60 days since last enrollment ended = 'plan_change' (excluded)
"""

from datetime import date, timedelta
from collections import defaultdict
from openpyxl import load_workbook


def parse_date(d) -> date | None:
    if not d:
        return None
    from datetime import datetime
    try:
        return datetime.strptime(str(d).strip(), "%m/%d/%Y").date()
    except Exception:
        return None


def parse_enrollment_report(xlsx_path: str, center: str, report_date: date = None) -> dict:
    """
    Parse the full enrollment Excel and return enrollment stats for the
    given center and reference date.

    Returns:
    {
        this_month:  [{name, start, grade, classification}, ...],
        last_month:  [...],
        this_week:   [...],
        plan_changes_this_month: [{name, start}, ...],
    }
    """
    if report_date is None:
        report_date = date.today()

    # ── Date boundaries ────────────────────────────────────────────────────────
    month_start = date(report_date.year, report_date.month, 1)
    if report_date.month == 1:
        last_month_start = date(report_date.year - 1, 12, 1)
    else:
        last_month_start = date(report_date.year, report_date.month - 1, 1)
    last_month_end = month_start - timedelta(days=1)
    week_start = report_date - timedelta(days=report_date.weekday())  # Monday

    # ── Load workbook ──────────────────────────────────────────────────────────
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = [
        dict(zip(headers, [cell.value for cell in row]))
        for row in ws.iter_rows(min_row=2)
    ]

    # ── Filter to this center ──────────────────────────────────────────────────
    center_rows = [r for r in rows if (r.get("Center") or "").strip() == center]

    # ── Filter dummy/test accounts ─────────────────────────────────────────────
    def is_dummy(r):
        account   = (r.get("Account Name") or "").lower()
        first     = (r.get("Student First Name") or "").lower()
        last      = (r.get("Student Last Name") or "").lower()
        guardians = (r.get("Guardians") or "").lower()
        emails    = (r.get("Guardian Emails") or "").lower()
        if "test" in first or "test" in last:
            return True
        if "matt diamond" in account or "matt diamond" in guardians:
            return True
        if "sample" in account or "sample" in first or "sample" in last:
            return True
        if "@mathnasium.com" in emails:
            return True
        return False

    center_rows = [r for r in center_rows if not is_dummy(r)]
    student_enrollments = defaultdict(list)
    for r in center_rows:
        start = parse_date(r.get("Primary Enrollment Start"))
        end   = parse_date(r.get("Primary Enrollment End"))
        if not start:
            continue
        # Key by (Lead Id, First, Last) so siblings on the same account
        # don't interfere with each other's classification
        key = (r["Lead Id"],
               (r.get("Student First Name") or "").strip(),
               (r.get("Student Last Name") or "").strip())
        student_enrollments[key].append({
            "name":  f"{r['Student First Name']} {r['Student Last Name']}".strip(),
            "grade": r.get("Grade") or "",
            "start": start,
            "end":   end,
            "type":  r.get("Membership Type") or "",
        })

    for key in student_enrollments:
        student_enrollments[key].sort(key=lambda x: x["start"])

    # ── Classify each enrollment ───────────────────────────────────────────────
    def classify(enrollments, idx):
        if idx == 0:
            return "new"
        prev = enrollments[idx - 1]
        curr = enrollments[idx]
        if prev["end"]:
            gap = (curr["start"] - prev["end"]).days
            if gap <= 60:
                return "plan_change"
        return "re-enrollment"

    classified = []
    for key, enrollments in student_enrollments.items():
        lead_id = key[0]
        for i, enr in enumerate(enrollments):
            classified.append({
                **enr,
                "lead_id":        lead_id,
                "classification": classify(enrollments, i),
            })

    # ── Separate plan changes from real enrollments ────────────────────────────
    real       = [c for c in classified if c["classification"] != "plan_change"]
    plan_chgs  = [c for c in classified if c["classification"] == "plan_change"]

    # ── Tag enrollment type ────────────────────────────────────────────────────
    def enroll_type(e):
        t = (e.get("type") or "").lower()
        if "private" in t: return "private"
        if "summer"  in t: return "summer"
        return "standard"

    for e in real:
        e["enroll_type"] = enroll_type(e)
    for e in plan_chgs:
        e["enroll_type"] = enroll_type(e)

    # ── Filter by date windows ─────────────────────────────────────────────────
    this_month  = [e for e in real if month_start      <= e["start"] <= report_date]
    last_month  = [e for e in real if last_month_start <= e["start"] <= last_month_end]
    this_week   = [e for e in real if week_start       <= e["start"] <= report_date]
    plan_this_month = [e for e in plan_chgs if month_start <= e["start"] <= report_date]

    # ── Total active roster ────────────────────────────────────────────────────
    # Count by student row (not Lead ID) since siblings share a Lead ID.
    # Use the most recent enrollment row per student name + lead combination.
    # A student is active if their current status is Enrolled or On Hold.
    from collections import defaultdict as _dd

    # Key by (Lead Id, First Name, Last Name) to handle siblings
    student_key_rows = _dd(list)
    for r in center_rows:
        key = (r["Lead Id"],
               (r.get("Student First Name") or "").strip(),
               (r.get("Student Last Name") or "").strip())
        start = parse_date(r.get("Primary Enrollment Start"))
        if start:
            student_key_rows[key].append({
                "start":  start,
                "status": (r.get("Status") or "").strip(),
                "type":   (r.get("Membership Type") or "").strip(),
            })

    enrolled_count = 0
    on_hold_count  = 0
    for key, records in student_key_rows.items():
        records.sort(key=lambda x: x["start"], reverse=True)
        latest = records[0]
        latest_status = latest["status"]
        latest_type   = latest.get("type", "").lower()
        # Exclude private memberships from roster count
        if "private" in latest_type:
            continue
        if latest_status == "Enrolled":
            enrolled_count += 1
        elif latest_status == "On Hold":
            on_hold_count += 1

    active = enrolled_count + on_hold_count

    def by_type(lst, t): return [e for e in lst if e.get("enroll_type") == t]

    return {
        "center":               center,
        "report_date":          report_date,
        "month_label":          report_date.strftime("%B %Y"),
        "last_month_label":     last_month_start.strftime("%B %Y"),
        "week_label":           f"Week of {week_start.strftime('%b %-d')}",
        "active_roster":        active,
        "enrolled_count":       enrolled_count,
        "on_hold_count":        on_hold_count,

        # All enrollments by window
        "this_month":           sorted(this_month,  key=lambda x: x["start"]),
        "last_month":           sorted(last_month,  key=lambda x: x["start"]),
        "this_week":            sorted(this_week,   key=lambda x: x["start"]),

        # Standard only
        "this_month_standard":  sorted(by_type(this_month, "standard"),  key=lambda x: x["start"]),
        "last_month_standard":  sorted(by_type(last_month, "standard"),  key=lambda x: x["start"]),
        "this_week_standard":   sorted(by_type(this_week,  "standard"),  key=lambda x: x["start"]),

        # Private only
        "this_month_private":   sorted(by_type(this_month, "private"),   key=lambda x: x["start"]),
        "last_month_private":   sorted(by_type(last_month, "private"),   key=lambda x: x["start"]),
        "this_week_private":    sorted(by_type(this_week,  "private"),   key=lambda x: x["start"]),

        # Summer only
        "this_month_summer":    sorted(by_type(this_month, "summer"),    key=lambda x: x["start"]),
        "last_month_summer":    sorted(by_type(last_month, "summer"),    key=lambda x: x["start"]),
        "this_week_summer":     sorted(by_type(this_week,  "summer"),    key=lambda x: x["start"]),

        "plan_changes_this_month": sorted(plan_this_month, key=lambda x: x["start"]),
    }


if __name__ == "__main__":
    import sys, json
    path   = sys.argv[1] if len(sys.argv) > 1 else "downloads/enrollment_today.xlsx"
    center = sys.argv[2] if len(sys.argv) > 2 else "Teaneck"
    data   = parse_enrollment_report(path, center)
    # Convert dates to strings for JSON output
    for key in ["this_month", "last_month", "this_week", "plan_changes_this_month"]:
        for e in data[key]:
            e["start"] = str(e["start"])
            if e.get("end"): e["end"] = str(e["end"])
    data["report_date"] = str(data["report_date"])
    print(json.dumps(data, indent=2))
