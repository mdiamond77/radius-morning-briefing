"""
parse.py
Reads the Radius Excel export and computes all stats needed for the report.
"""

import re
from collections import defaultdict
from datetime import date
from openpyxl import load_workbook


def to_minutes(time_str) -> int | None:
    if not time_str or not str(time_str).strip():
        return None
    try:
        t = str(time_str).strip()
        parts = t.split(" ")
        h, m = map(int, parts[0].split(":"))
        if len(parts) > 1 and parts[1].upper() == "PM" and h != 12:
            h += 12
        if len(parts) > 1 and parts[1].upper() == "AM" and h == 12:
            h = 0
        return h * 60 + m
    except Exception:
        return None


def fmt_time(minutes: int) -> str:
    h = minutes // 60
    m = minutes % 60
    ampm = "PM" if h >= 12 else "AM"
    h12 = h - 12 if h > 12 else (12 if h == 0 else h)
    return f"{h12}:{m:02d} {ampm}"


def parse_lp(lp_str: str) -> tuple[list[str], list[str]]:
    if not lp_str:
        return [], []
    mastered, worked = [], []
    for entry in str(lp_str).split(";"):
        entry = entry.strip()
        if not entry:
            continue
        match = re.search(r'\(([^)]+)\)', entry)
        topic = match.group(1).strip() if match else entry
        if "Mastered" in entry:
            mastered.append(topic)
        elif "Worked On" in entry:
            worked.append(topic)
    return mastered, worked


def half_hour_buckets(sessions: list[dict], center: str = "") -> list[dict]:
    """Build half-hour buckets with student count, instructor count, and ratio.
    Center directors are excluded from instructor counts."""

    # Center directors — excluded from ratio calculations
    CENTER_DIRECTORS = {"elizabeth anacleto", "samba taha"}

    all_starts = [s["start_min"] for s in sessions if s["start_min"]]
    all_ends   = [s["end_min"]   for s in sessions if s["end_min"]]
    if not all_starts:
        return []

    earliest = (min(all_starts) // 30) * 30
    latest   = ((max(all_ends) + 29) // 30) * 30

    buckets = []
    for bucket_start in range(earliest, latest, 30):
        bucket_end = bucket_start + 30

        students_in    = set()
        instructors_in = set()

        for s in sessions:
            if s["start_min"] and s["end_min"] \
               and s["start_min"] < bucket_end \
               and s["end_min"] > bucket_start:
                students_in.add(s["name"])
                for inst in [i.strip() for i in s["instructor"].split(",") if i.strip()]:
                    if inst.lower() not in CENTER_DIRECTORS:
                        instructors_in.add(inst)

        if students_in:
            sc = len(students_in)
            ic = len(instructors_in)
            ratio = round(sc / ic, 1) if ic else None
            buckets.append({
                "label":       f"{fmt_time(bucket_start)} \u2013 {fmt_time(bucket_end)}",
                "count":       sc,
                "instructors": ic,
                "ratio":       ratio,
                "students":    sorted(students_in),
                "peak":        False,
            })

    if buckets:
        max_count = max(b["count"] for b in buckets)
        for b in buckets:
            if b["count"] == max_count:
                b["peak"] = True
                break

    return buckets


def parse_report(xlsx_path: str, report_date: date = None, private_students: set = None) -> dict:
    if report_date is None:
        report_date = date.today()
    if private_students is None:
        private_students = set()

    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]

    if not rows:
        return {}

    sessions = []
    for row in rows:
        mastered, worked_on = parse_lp(row.get("LP Assignment") or "")
        start_min = to_minutes(row.get("Session Start"))
        end_min   = to_minutes(row.get("Session End"))

        try:
            pages = int(row.get("Pages Completed") or 0)
        except (ValueError, TypeError):
            pages = 0

        try:
            goal = int(row.get("Session Page Goal") or 0)
        except (ValueError, TypeError):
            goal = 0

        try:
            score_raw = str(row.get("Mathlete Score") or "").strip()
            score = float(score_raw) if score_raw else None
        except (ValueError, TypeError):
            score = None

        sessions.append({
            "name":                   str(row.get("Student Name") or "").strip(),
            "instructor":             str(row.get("Instructors") or "").strip(),
            "start":                  str(row.get("Session Start") or "").strip(),
            "end":                    str(row.get("Session End") or "").strip(),
            "start_min":              start_min,
            "end_min":                end_min,
            "pages":                  pages,
            "goal":                   goal,
            "beat_goal":              pages > goal if goal else False,
            "score":                  score,
            "mastered":               mastered,
            "worked_on":              worked_on,
            "session_notes":          str(row.get("Session Summary Notes") or "").strip(),
            "internal_notes":         str(row.get("Internal Notes") or "").strip(),
            "schoolwork_description": str(row.get("Schoolwork Description") or "").strip(),
            "membership_type":        str(row.get("Membership Type") or "").strip(),
            "lp_assigned":            bool(str(row.get("LP Assignment") or "").strip()),
            "assessment":             str(row.get("Assessment") or "").strip(),
            "delivery":               str(row.get("Delivery Method") or "").strip(),
            "center":                 str(row.get("Center") or "").strip(),
        })

    total_sessions  = len(sessions)
    unique_students = len(set(s["name"] for s in sessions))
    total_pages     = sum(s["pages"] for s in sessions)

    scores    = [s["score"] for s in sessions if s["score"] is not None]
    avg_score = round(sum(scores) / len(scores), 1) if scores else None

    mastery_list  = [{"name": s["name"], "topics": s["mastered"]} for s in sessions if s["mastered"]]
    total_mastery = sum(len(m["topics"]) for m in mastery_list)

    CENTER_DIRECTORS = {"elizabeth anacleto", "samba taha"}

    instructor_map = defaultdict(list)
    team_taught    = set()
    for s in sessions:
        instructors = [i.strip() for i in s["instructor"].split(",") if i.strip()]
        for inst in instructors:
            instructor_map[inst].append(s["name"])
        if len(instructors) > 1:
            team_taught.add(s["name"])

    instructor_summary = []
    for inst, students in sorted(instructor_map.items(), key=lambda x: -len(x[1])):
        solo   = [st for st in students if st not in team_taught]
        teamed = [st for st in students if st in team_taught]
        parts  = []
        if solo:   parts.append(f"{len(solo)} solo")
        if teamed: parts.append(f"{len(teamed)} team-taught")
        instructor_summary.append({
            "name":               inst,
            "count":              len(students),
            "students":           students,
            "detail":             ", ".join(parts),
            "is_center_director": inst.lower() in CENTER_DIRECTORS,
        })

    # ── Assessment buckets ─────────────────────────────────────────────────────
    def in_bucket(s, keyword):
        a = s["assessment"].lower()
        return keyword in a

    assessments = {
        "pre_in_progress":        sorted({s["name"] for s in sessions if in_bucket(s, "pre in progress")}),
        "pre_completed":          sorted({s["name"] for s in sessions if in_bucket(s, "pre") and in_bucket(s, "completed") and not in_bucket(s, "progress check")}),
        "post_in_progress":       sorted({s["name"] for s in sessions if in_bucket(s, "post in progress")}),
        "post_completed":         sorted({s["name"] for s in sessions if in_bucket(s, "post") and in_bucket(s, "completed") and not in_bucket(s, "progress check")}),
        "progress_in_progress":   sorted({s["name"] for s in sessions if in_bucket(s, "progress check in progress")}),
        "progress_completed":     sorted({s["name"] for s in sessions if in_bucket(s, "progress check completed")}),
    }
    assessment_students = {s["name"] for s in sessions if s["assessment"]}

    # ── Missing LP logic ───────────────────────────────────────────────────────
    STUDY_KEYWORDS = [
        "homework", "hw", "test", "quiz", "exam", "review", "study",
        "studying", "school work", "schoolwork",
    ]

    def is_private(s):
        return s["name"] in private_students

    def looks_like_study(s):
        combined = " ".join([
            s.get("session_notes") or "",
            s.get("schoolwork_description") or "",
        ]).lower()
        return any(kw in combined for kw in STUDY_KEYWORDS)

    missing_lp_flagged   = []  # genuinely missing — show in red
    missing_lp_study     = []  # likely study/hw — show with note

    for s in sessions:
        if s["lp_assigned"]:
            continue
        if s["name"] in assessment_students:
            continue  # assessment explains missing LP
        if looks_like_study(s):
            missing_lp_study.append(s["name"])
        else:
            missing_lp_flagged.append(s["name"])

    private_count = sum(1 for s in sessions if is_private(s))

    return {
        "report_date":        report_date.strftime("%A, %B %-d, %Y"),
        "center":             sessions[0]["center"].split(",")[0].strip() if sessions else "Center",
        "total_sessions":     total_sessions,
        "private_sessions":   private_count,
        "unique_students":    unique_students,
        "total_pages":        total_pages,
        "avg_score":          avg_score,
        "total_mastery":      total_mastery,
        "mastery_list":       mastery_list,
        "instructor_summary": instructor_summary,
        "att_buckets":        half_hour_buckets(sessions),
        "internal_notes":     [{"name": s["name"], "note": s["internal_notes"]} for s in sessions if s["internal_notes"]],
        "missing_lp":         missing_lp_flagged,
        "missing_lp_study":   missing_lp_study,
        "beat_goal":          [s["name"] for s in sessions if s["beat_goal"]],
        "below_goal":         [f"{s['name']} ({s['pages']}/{s['goal']} pages)" for s in sessions if not s["beat_goal"] and s["goal"] > 0],
        "assessments":        assessments,
        "sessions":           sessions,
    }


if __name__ == "__main__":
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else "downloads/radius_today.xlsx"
    data = parse_report(path)
    print(json.dumps(data, indent=2, default=str))
